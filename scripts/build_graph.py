#!/usr/bin/env python3
import argparse
import csv
import json
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


def sniff_dialect(path: Path) -> csv.Dialect:
    with path.open("r", encoding="utf-8") as handle:
        sample = handle.read(2048)
    try:
        return csv.Sniffer().sniff(sample, delimiters=";,	")
    except csv.Error:
        return csv.excel


def load_stop_parent_map(stops_path: Path) -> dict:
    if not stops_path.exists():
        return {}
    if stops_path.suffix.lower() == ".xlsx":
        workbook = load_workbook(stops_path, read_only=True, data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        rows = worksheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            return {}
        header = [str(value).strip() for value in header_row if value is not None]
        col = {name: idx for idx, name in enumerate(header)}
        stop_id_idx = col.get("stop_id")
        parent_idx = col.get("parent_station")
        location_idx = col.get("location_type")
        if stop_id_idx is None:
            return {}
        mapping = {}
        for row in rows:
            if not row or stop_id_idx >= len(row):
                continue
            stop_id = row[stop_id_idx]
            if not stop_id:
                continue
            parent = None
            if parent_idx is not None and parent_idx < len(row):
                parent = row[parent_idx]
            if location_idx is not None and location_idx < len(row):
                if row[location_idx] == 1:
                    parent = stop_id
            mapping[str(stop_id).strip()] = str(parent).strip() if parent else None
        return mapping

    mapping = {}
    dialect = sniff_dialect(stops_path)
    with stops_path.open("r", encoding="utf-8") as handle:
        reader = csv.DictReader(handle, dialect=dialect)
        if not reader.fieldnames:
            return mapping
        for row in reader:
            stop_id = row.get("stop_id")
            if not stop_id:
                continue
            parent = row.get("parent_station")
            location = row.get("location_type")
            if location in ("1", 1):
                parent = stop_id
            mapping[str(stop_id).strip()] = str(parent).strip() if parent else None
    return mapping


def main() -> int:
    parser = argparse.ArgumentParser(description="Build a stop graph from GTFS stop_times.")
    parser.add_argument("--stop-times", type=Path, required=True)
    parser.add_argument("--output", type=Path, default=Path("data/graph.json"))
    parser.add_argument("--stops", type=Path, default=None)
    parser.add_argument("--limit-trips", type=int, default=None)
    args = parser.parse_args()

    if not args.stop_times.exists():
        return 1

    dialect = sniff_dialect(args.stop_times)
    trips = defaultdict(list)
    parent_map = load_stop_parent_map(args.stops) if args.stops else {}

    with args.stop_times.open("r", encoding="utf-8") as handle:
        reader = csv.DictReader(handle, dialect=dialect)
        if not reader.fieldnames:
            return 1
        if "trip_id" not in reader.fieldnames or "stop_id" not in reader.fieldnames:
            return 1
        sequence_field = "stop_sequence" if "stop_sequence" in reader.fieldnames else None
        for row in reader:
            trip_id = row.get("trip_id")
            stop_id = row.get("stop_id")
            if not trip_id or not stop_id:
                continue
            if parent_map:
                stop_id = parent_map.get(stop_id, stop_id)
            if sequence_field:
                try:
                    seq = int(row.get(sequence_field, 0))
                except ValueError:
                    seq = 0
            else:
                seq = len(trips[trip_id])
            trips[trip_id].append((seq, stop_id))
            if args.limit_trips and len(trips) >= args.limit_trips:
                break

    edges = defaultdict(set)
    for trip_id, stops in trips.items():
        ordered = [stop_id for _, stop_id in sorted(stops, key=lambda item: item[0])]
        for a, b in zip(ordered, ordered[1:]):
            if a == b:
                continue
            edges[a].add(b)
            edges[b].add(a)

    graph = {
        "edges": {node: sorted(list(neighbors)) for node, neighbors in edges.items()},
        "meta": {
            "node_count": len(edges),
            "edge_count": sum(len(neigh) for neigh in edges.values()),
        },
    }

    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w", encoding="utf-8") as handle:
        json.dump(graph, handle, ensure_ascii=True, indent=2)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

#!/usr/bin/env python3
import argparse
import csv
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.append(str(ROOT))

from src.travel_order_resolver import normalize


def read_rows(path: Path) -> tuple[list[str], list[list]]:
    if path.suffix.lower() == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return [], []
        header = [str(value).strip() for value in rows[0] if value is not None]
        return header, rows[1:]

    with path.open("r", encoding="utf-8") as handle:
        sample = handle.read(2048)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,	")
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(handle, dialect=dialect)
        rows = list(reader)
        if not rows:
            return [], []
        header = [value.strip() for value in rows[0]]
        return header, rows[1:]


def main() -> int:
    parser = argparse.ArgumentParser(description="Build stop area index from stops data.")
    parser.add_argument("--input", type=Path, default=Path("stops.xlsx"))
    parser.add_argument("--output-csv", type=Path, default=Path("data/stops_areas.csv"))
    parser.add_argument("--output-json", type=Path, default=Path("data/stops_index.json"))
    parser.add_argument("--limit", type=int, default=None)
    args = parser.parse_args()

    if not args.input.exists():
        return 1

    header, rows = read_rows(args.input)
    if not header:
        return 1

    col = {name: idx for idx, name in enumerate(header)}
    required = ["stop_id", "stop_name"]
    for req in required:
        if req not in col:
            return 1

    location_idx = col.get("location_type")

    areas = []
    for row in rows:
        if args.limit and len(areas) >= args.limit:
            break
        if location_idx is not None and location_idx < len(row):
            location = row[location_idx]
            if location not in (1, "1", None):
                continue
        stop_id = row[col["stop_id"]]
        stop_name = row[col["stop_name"]]
        if not stop_id or not stop_name:
            continue
        areas.append((str(stop_id).strip(), str(stop_name).strip()))

    args.output_csv.parent.mkdir(parents=True, exist_ok=True)
    with args.output_csv.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["stop_id", "stop_name", "normalized"])
        for stop_id, stop_name in areas:
            writer.writerow([stop_id, stop_name, normalize(stop_name)])

    index = {}
    for stop_id, stop_name in areas:
        key = normalize(stop_name)
        entry = index.setdefault(key, {"names": set(), "stop_ids": set()})
        entry["names"].add(stop_name)
        entry["stop_ids"].add(stop_id)

    for key in index:
        index[key]["names"] = sorted(index[key]["names"])
        index[key]["stop_ids"] = sorted(index[key]["stop_ids"])

    args.output_json.parent.mkdir(parents=True, exist_ok=True)
    with args.output_json.open("w", encoding="utf-8") as handle:
        json.dump(index, handle, ensure_ascii=True, indent=2)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

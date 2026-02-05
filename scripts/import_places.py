#!/usr/bin/env python3
import argparse
import csv
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

DEFAULT_COLUMNS = [
    "name",
    "nom",
    "libelle",
    "label",
    "station_name",
    "stop_name",
    "gare",
]


def select_column(header: list[str], preferred: str | None) -> str | None:
    lowered = {col.lower(): col for col in header}
    if preferred:
        return lowered.get(preferred.lower())
    for candidate in DEFAULT_COLUMNS:
        if candidate in lowered:
            return lowered[candidate]
    return None


def extract_from_csv(path: Path, preferred_column: str | None) -> list[str]:
    names = []
    with path.open("r", encoding="utf-8") as handle:
        sample = handle.read(2048)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
        except csv.Error:
            dialect = csv.excel
        reader = csv.DictReader(handle, dialect=dialect)
        if not reader.fieldnames:
            return names
        column = select_column(reader.fieldnames, preferred_column)
        if not column:
            return names
        for row in reader:
            value = row.get(column, "").strip()
            if value:
                names.append(value)
    return names


def normalize_header(values: Iterable[str]) -> list[str]:
    return [str(value).strip() for value in values if value is not None]


def extract_from_xlsx(path: Path, preferred_column: str | None) -> list[str]:
    names = []
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    header_row = next(rows, None)
    if not header_row:
        return names
    header = normalize_header(header_row)
    column = select_column(header, preferred_column)
    if not column:
        return names
    column_index = header.index(column)
    location_index = header.index("location_type") if "location_type" in header else None
    for row in rows:
        if not row or column_index >= len(row):
            continue
        if location_index is not None and location_index < len(row):
            if row[location_index] not in (1, "1", None):
                continue
        value = row[column_index]
        if value is None:
            continue
        text = str(value).strip()
        if text:
            names.append(text)
    return names


def main() -> int:
    parser = argparse.ArgumentParser(description="Import place names from CSV.")
    parser.add_argument("--input", type=Path, required=True)
    parser.add_argument("--output", type=Path, default=Path("data/places_imported.txt"))
    parser.add_argument("--column", type=str, default=None)
    parser.add_argument("--add-gare-alias", action="store_true")
    parser.add_argument("--limit", type=int, default=None)
    args = parser.parse_args()

    if not args.input.exists():
        return 1

    suffix = args.input.suffix.lower()
    if suffix == ".xlsx":
        names = extract_from_xlsx(args.input, args.column)
    else:
        names = extract_from_csv(args.input, args.column)
    if not names:
        return 1
    if args.limit:
        names = names[: args.limit]

    unique = []
    seen = set()
    for name in names:
        if name in seen:
            continue
        unique.append(name)
        seen.add(name)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w", encoding="utf-8") as handle:
        for name in unique:
            handle.write(name + "\n")
            if args.add_gare_alias:
                handle.write(f"Gare de {name}|{name}\n")
                handle.write(f"Gare {name}|{name}\n")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
